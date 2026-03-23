"""
Excel export service — maandplanning in v0.7 HR-formaat.

Identiek formaat als v0.7 ExportService:
  Sheet 1 "Blad1"          : diensttabel (kolom A = maand, B = naam, C+ = shifts)
  Sheet 2 "Validatie Rapport": HR-overtredingen met samenvatting
"""
import logging
from calendar import monthrange
from datetime import date, datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from models.gebruiker import Gebruiker
from models.lidmaatschap import Lidmaatschap
from models.planning import Planning
from services.domein.balans_domein import belgische_feestdagen
from services.domein.planning_domein import MAAND_NAMEN
from services.domein.validatie_domein import ValidatieFout

logger = logging.getLogger(__name__)

# ──────────────────────────────────────────────
# Kleurconstanten — exact overgenomen uit v0.7
# ──────────────────────────────────────────────
_KLEUREN = {
    "maand_header": "FFDDEBF7",   # lichtblauw — maandnaam cel (kolom A)
    "datum_header": "FFBDD7EE",   # iets donkerder blauw — datumkop + naam-kop
    "naam_kolom":   "FFD3D3D3",   # lichtgrijs — "Naam + Voornaam" kop
    "feestdag":     "FFF0E68C",   # geel — feestdagen
    "weekend":      "FFF8CBAD",   # oranje — za/zo
    "werkdag":      "FFBFBFBF",   # grijs — ma–vr header
}
_HEADER_FILL  = PatternFill(start_color="FF366092", end_color="FF366092", fill_type="solid")
_HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
_TITEL_FONT   = Font(bold=True, size=14, color="366092")
_GROEN_FILL   = PatternFill(start_color="FFC8E6C9", end_color="FFC8E6C9", fill_type="solid")
_ORANJE_FILL  = PatternFill(start_color="FFFFB74D", end_color="FFFFB74D", fill_type="solid")
_ROOD_FILL    = PatternFill(start_color="FFFFCDD2", end_color="FFFFCDD2", fill_type="solid")
_BORDER_THIN  = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)
_BORDER_MEDIUM_LEFT  = Side(style="medium")
_BORDER_MEDIUM_RIGHT = Side(style="medium")
_BORDER_MEDIUM_TOP   = Side(style="medium")

# Korte maandnamen (voor "jan/26" header cel)
_MAAND_KORT = {
    1: "jan", 2: "feb", 3: "mrt", 4: "apr",  5: "mei",  6: "jun",
    7: "jul", 8: "aug", 9: "sep", 10: "okt", 11: "nov", 12: "dec",
}


def _fill(argb: str) -> PatternFill:
    return PatternFill(start_color=argb, end_color=argb, fill_type="solid")


class ExcelExportService:
    def __init__(self, db: Session) -> None:
        self.db = db

    def genereer_excel(
        self,
        team_id: int,
        jaar: int,
        maand: int,
        fouten: list[ValidatieFout] | None = None,
    ) -> bytes:
        """
        Genereer een .xlsx bestand in v0.7 HR-formaat.

        Args:
            team_id: Team van de ingelogde planner.
            jaar: Jaar van de te exporteren maand.
            maand: Maand (1–12) van de te exporteren maand.
            fouten: HR-validatiefouten voor Sheet 2 (optioneel).

        Returns:
            Ruwe bytes van het .xlsx bestand.
        """
        import io

        _, aantal_dagen = monthrange(jaar, maand)
        datums = [date(jaar, maand, d) for d in range(1, aantal_dagen + 1)]
        feestdagen = belgische_feestdagen(jaar)

        gebruikers = (
            self.db.query(Gebruiker)
            .join(Lidmaatschap, Lidmaatschap.gebruiker_id == Gebruiker.id)
            .filter(
                Lidmaatschap.team_id == team_id,
                Lidmaatschap.is_actief == True,
                Lidmaatschap.verwijderd_op == None,
                Gebruiker.is_actief == True,
            )
            .order_by(Gebruiker.volledige_naam)
            .all()
        )
        shifts_db = (
            self.db.query(Planning)
            .filter(
                Planning.team_id == team_id,
                Planning.datum >= datums[0],
                Planning.datum <= datums[-1],
            )
            .all()
        )
        shifts_idx: dict[tuple[int, date], str | None] = {
            (s.gebruiker_id, s.datum): s.shift_code for s in shifts_db
        }

        # Bouw planning_data in zelfde structuur als v0.7
        planning_data = [
            {
                "naam": g.volledige_naam or g.gebruikersnaam,
                "planning": {
                    d.strftime("%Y-%m-%d"): shifts_idx.get((g.id, d), "") or ""
                    for d in datums
                },
            }
            for g in gebruikers
        ]

        wb = Workbook()
        ws = wb.active
        ws.title = "Blad1"
        self._vul_diensttabel(ws, jaar, maand, datums, feestdagen, planning_data)

        ws2 = wb.create_sheet(title="Validatie Rapport")
        self._vul_validatie_rapport(ws2, jaar, maand, fouten or [])

        buf = io.BytesIO()
        wb.save(buf)
        logger.debug(
            "Excel gegenereerd voor team %s %s-%02d: %d medewerkers",
            team_id, jaar, maand, len(gebruikers),
        )
        return buf.getvalue()

    # ------------------------------------------------------------------ #
    # Sheet 1 — Diensttabel                                               #
    # ------------------------------------------------------------------ #

    def _vul_diensttabel(
        self,
        ws,
        jaar: int,
        maand: int,
        datums: list[date],
        feestdagen: frozenset[date],
        planning_data: list[dict],
    ) -> None:
        dagen_in_maand = len(datums)
        maand_naam = MAAND_NAMEN[maand]
        jaar_kort = str(jaar)[-2:]
        center = Alignment(horizontal="center", vertical="center")
        left_top = Alignment(horizontal="left", vertical="top", wrap_text=True)

        # ── Rij 2: maandnaam (A) | mmm/jj (B) | beschrijving (C, gemerged) ──
        a2 = ws["A2"]
        a2.value = maand_naam.upper()
        a2.font = Font(name="Arial", size=14, bold=True)
        a2.fill = _fill(_KLEUREN["maand_header"])
        a2.alignment = center
        a2.border = Border(
            left=Side(style="medium"), right=Side(style="medium"), top=Side(style="medium")
        )

        b2 = ws["B2"]
        b2.value = f"{_MAAND_KORT[maand]}/{jaar_kort}"
        b2.font = Font(name="Arial", size=11, bold=True)
        b2.fill = _fill(_KLEUREN["datum_header"])
        b2.alignment = center
        b2.border = Border(left=Side(style="medium"), top=Side(style="medium"))

        c2 = ws["C2"]
        c2.value = (
            f"Diensttabel van {datums[0].strftime('%d-%m-%Y')} "
            f"tot {datums[-1].strftime('%d-%m-%Y')}"
        )
        c2.font = Font(name="Arial", size=11, bold=True)
        c2.fill = _fill(_KLEUREN["datum_header"])
        c2.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c2.border = Border(
            left=Side(style="medium"), top=Side(style="medium"), bottom=Side(style="medium")
        )
        last_col = get_column_letter(2 + dagen_in_maand)
        ws.merge_cells(f"C2:{last_col}2")

        # ── Rij 3: "Naam + Voornaam" (B) | dagnummers met kleuren (C+) ──
        b3 = ws["B3"]
        b3.value = "Naam + Voornaam"
        b3.font = Font(name="Arial", size=10, bold=True)
        b3.fill = _fill(_KLEUREN["naam_kolom"])
        b3.alignment = left_top
        b3.border = Border(
            left=Side(style="medium"), right=Side(style="thin"),
            top=Side(style="medium"), bottom=Side(style="medium"),
        )

        for dag_nr, d in enumerate(datums, start=1):
            col_idx = 2 + dag_nr
            is_feestdag = d in feestdagen
            is_weekend = d.weekday() >= 5
            if is_feestdag:
                kleur = _KLEUREN["feestdag"]
            elif is_weekend:
                kleur = _KLEUREN["weekend"]
            else:
                kleur = _KLEUREN["werkdag"]

            is_eerste = col_idx == 3
            is_laatste = dag_nr == dagen_in_maand
            cel = ws.cell(row=3, column=col_idx)
            cel.value = dag_nr
            cel.font = Font(name="Arial", size=10, bold=True)
            cel.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cel.fill = _fill(kleur)
            cel.border = Border(
                left=Side(style="medium" if is_eerste else "thin"),
                right=Side(style="medium" if is_laatste else "thin"),
                top=Side(style="medium"),
            )

        # ── Datarijen (rij 4+) ──
        for rij_offset, gd in enumerate(planning_data):
            rij = 4 + rij_offset
            is_eerste_rij = rij_offset == 0

            naam_cel = ws.cell(row=rij, column=2)
            naam_cel.value = gd["naam"]
            naam_cel.font = Font(name="Arial", size=9, bold=True)
            naam_cel.alignment = left_top
            naam_cel.border = Border(
                left=Side(style="medium"),
                top=Side(style="medium" if is_eerste_rij else "thin"),
                bottom=Side(style="thin"),
            )

            for dag_nr, d in enumerate(datums, start=1):
                col_idx = 2 + dag_nr
                datum_str = d.strftime("%Y-%m-%d")
                shift_code = gd["planning"].get(datum_str, "")
                is_feestdag = d in feestdagen
                is_weekend = d.weekday() >= 5
                is_eerste_col = col_idx == 3
                is_laatste_col = dag_nr == dagen_in_maand

                cel = ws.cell(row=rij, column=col_idx)
                cel.value = shift_code
                cel.font = Font(name="Arial", size=10)
                cel.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                if is_feestdag:
                    cel.fill = _fill(_KLEUREN["feestdag"])
                elif is_weekend:
                    cel.fill = _fill(_KLEUREN["weekend"])
                cel.border = Border(
                    left=Side(style="medium" if is_eerste_col else "thin"),
                    right=Side(style="medium" if is_laatste_col else "thin"),
                    top=Side(style="medium" if is_eerste_rij else "thin"),
                    bottom=Side(style="thin"),
                )

        # ── Kolom A verticaal gemerged (maandnaam) ──
        laatste_rij = 3 + len(planning_data)
        if laatste_rij >= 2:
            ws.merge_cells(f"A2:A{laatste_rij}")

        # ── Kolombreedte en rijhoogte ──
        for col_idx in range(1, 3 + dagen_in_maand):
            ws.column_dimensions[get_column_letter(col_idx)].width = 13.0
        ws.row_dimensions[2].height = 15.75
        ws.row_dimensions[3].height = 39.0
        for rij in range(4, 4 + len(planning_data)):
            ws.row_dimensions[rij].height = 25.0

    # ------------------------------------------------------------------ #
    # Sheet 2 — Validatie Rapport                                         #
    # ------------------------------------------------------------------ #

    def _vul_validatie_rapport(
        self,
        ws,
        jaar: int,
        maand: int,
        fouten: list[ValidatieFout],
    ) -> None:
        maand_naam = MAAND_NAMEN[maand]
        center = Alignment(horizontal="center", vertical="center")
        left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)

        # Titel
        ws.merge_cells("A1:E1")
        titel = ws["A1"]
        titel.value = f"HR Validatie Rapport — {maand_naam.capitalize()} {jaar}"
        titel.font = _TITEL_FONT
        titel.alignment = center

        # Header
        for col, h in enumerate(["Datum", "Medewerker", "Ernst", "Regel", "Bericht"], start=1):
            cel = ws.cell(row=2, column=col)
            cel.value = h
            cel.font = _HEADER_FONT
            cel.fill = _HEADER_FILL
            cel.border = _BORDER_THIN
            cel.alignment = center

        # Datarijen
        for rij_nr, f in enumerate(fouten, start=3):
            ws.cell(row=rij_nr, column=1, value=f.datum.strftime("%d-%m-%Y")).border = _BORDER_THIN
            ws.cell(row=rij_nr, column=1).alignment = center

            mw_cel = ws.cell(row=rij_nr, column=2)
            mw_cel.value = f.gebruiker_naam if f.gebruiker_id != 0 else "(Dag-niveau)"
            mw_cel.border = _BORDER_THIN
            mw_cel.alignment = left_wrap

            ernst_cel = ws.cell(row=rij_nr, column=3)
            if f.ernst == "INFO":
                ernst_cel.value = "Info"
                ernst_cel.fill = _GROEN_FILL
            elif f.ernst == "WARNING":
                ernst_cel.value = "Waarschuwing"
                ernst_cel.fill = _ORANJE_FILL
            else:
                ernst_cel.value = "Kritiek"
                ernst_cel.fill = _ROOD_FILL
            ernst_cel.border = _BORDER_THIN
            ernst_cel.alignment = center

            regel_cel = ws.cell(row=rij_nr, column=4, value=f.validator_code)
            regel_cel.border = _BORDER_THIN
            regel_cel.alignment = left_wrap

            bericht_cel = ws.cell(row=rij_nr, column=5, value=f.bericht)
            bericht_cel.border = _BORDER_THIN
            bericht_cel.alignment = left_wrap

        # Samenvatting
        sam_start = len(fouten) + 5
        ws.merge_cells(f"A{sam_start}:E{sam_start}")
        ws.cell(row=sam_start, column=1, value="-" * 60)
        ws.cell(row=sam_start + 1, column=1, value="TOTAAL OVERZICHT").font = Font(bold=True, size=12)

        n_kritiek   = sum(1 for f in fouten if f.ernst == "CRITICAL")
        n_warning   = sum(1 for f in fouten if f.ernst == "WARNING")
        n_info      = sum(1 for f in fouten if f.ernst == "INFO")
        items = [
            ("Kritieke fouten:",  str(n_kritiek),         _ROOD_FILL),
            ("Waarschuwingen:",   str(n_warning),         _ORANJE_FILL),
            ("Informatief:",      str(n_info),            _GROEN_FILL),
            ("TOTAAL:",           str(len(fouten)),       None),
        ]
        for i, (label, waarde, fill) in enumerate(items):
            rij = sam_start + 3 + i
            lbl = ws.cell(row=rij, column=1, value=label)
            lbl.font = Font(bold=True, size=10)
            val = ws.cell(row=rij, column=2, value=waarde)
            val.font = Font(size=10)
            if fill:
                lbl.fill = fill
                val.fill = fill

        # Kolombreedte
        for col, breedte in enumerate([15, 25, 15, 25, 50], start=1):
            ws.column_dimensions[get_column_letter(col)].width = breedte
        ws.row_dimensions[1].height = 25
        ws.row_dimensions[2].height = 30
